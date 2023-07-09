from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
EXCEL_FILE_PATH = 'data.xlsx'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit_form():
    name = request.form.get('name')
    place = request.form.get('place')
    age = request.form.get('age')

    # Load existing workbook or create a new one
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Place', 'Age'])

    # Append the form data to the sheet
    sheet.append([name, place, age])

    # Save the workbook
    workbook.save(EXCEL_FILE_PATH)

    return redirect('http://localhost:5500/')  # Update with your forms app URL

if __name__ == '__main__':
    app.run()
